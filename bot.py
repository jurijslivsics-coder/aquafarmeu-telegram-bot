import asyncio
import json
import logging
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import httpx
from dotenv import load_dotenv
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook
from openai import AsyncOpenAI
from pypdf import PdfReader
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, filters


load_dotenv()

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("aquafarmeu-bot")

ROOT = Path(__file__).parent
KNOWLEDGE_DIR = ROOT / "knowledge"
URLS_PATH = KNOWLEDGE_DIR / "urls.txt"
INDEX_PATH = ROOT / "knowledge_index.json"

TELEGRAM_BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]
OPENAI_CHAT_MODEL = os.getenv("OPENAI_CHAT_MODEL", "gpt-4.1-mini")
OPENAI_EMBEDDING_MODEL = os.getenv("OPENAI_EMBEDDING_MODEL", "text-embedding-3-small")
WEBHOOK_URL = os.getenv("WEBHOOK_URL", "").rstrip("/")
PORT = int(os.getenv("PORT", "10000"))


def parse_allowed_users() -> set[int]:
    raw = os.getenv("ALLOWED_TELEGRAM_USER_IDS", "").strip()
    if not raw:
        return set()
    return {int(item.strip()) for item in raw.split(",") if item.strip()}


ALLOWED_USER_IDS = parse_allowed_users()
client = AsyncOpenAI(api_key=OPENAI_API_KEY)


@dataclass
class Chunk:
    source: str
    text: str
    embedding: list[float]


class KnowledgeBase:
    def __init__(self) -> None:
        self.chunks: list[Chunk] = []
        self.loaded = False
        self.loading_lock = asyncio.Lock()

    async def load(self) -> None:
        async with self.loading_lock:
            files = list(find_knowledge_files())
            if not files and not read_urls():
                logger.warning("No knowledge files found in %s", KNOWLEDGE_DIR)
                self.chunks = []
                self.loaded = True
                return

            if self._load_cached_index(files):
                logger.info("Loaded %s chunks from cache", len(self.chunks))
                self.loaded = True
                return

            text_chunks: list[tuple[str, str]] = []
            for file_path in files:
                logger.info("Reading knowledge file: %s", file_path.name)
                text = extract_text(file_path)
                for chunk in split_text(text):
                    text_chunks.append((file_path.name, chunk))

            for url in read_urls():
                try:
                    logger.info("Reading knowledge URL: %s", url)
                    text = await fetch_url_text(url)
                except Exception:
                    logger.exception("Could not fetch URL: %s", url)
                    continue
                for chunk in split_text(text):
                    text_chunks.append((url, chunk))

            if not text_chunks:
                self.chunks = []
                self.loaded = True
                return

            logger.info("Embedding %s knowledge chunks", len(text_chunks))
            embeddings = await embed_texts([text for _, text in text_chunks])
            self.chunks = [
                Chunk(source=source, text=text, embedding=embedding)
                for (source, text), embedding in zip(text_chunks, embeddings)
            ]
            self._save_index(files)
            self.loaded = True
            logger.info("Built knowledge index with %s chunks", len(self.chunks))

    def search(self, query_embedding: list[float], limit: int = 5) -> list[Chunk]:
        if not self.chunks:
            return []

        query = np.array(query_embedding)
        scored: list[tuple[float, Chunk]] = []
        for chunk in self.chunks:
            candidate = np.array(chunk.embedding)
            score = float(np.dot(query, candidate) / (np.linalg.norm(query) * np.linalg.norm(candidate)))
            scored.append((score, chunk))

        scored.sort(key=lambda item: item[0], reverse=True)
        return [chunk for score, chunk in scored[:limit] if score > 0.2]

    def _load_cached_index(self, files: list[Path]) -> bool:
        if not INDEX_PATH.exists():
            return False

        try:
            payload = json.loads(INDEX_PATH.read_text(encoding="utf-8"))
            expected = file_fingerprint(files)
            if payload.get("fingerprint") != expected:
                return False

            self.chunks = [Chunk(**item) for item in payload["chunks"]]
            return True
        except Exception:
            logger.exception("Could not load cached knowledge index")
            return False

    def _save_index(self, files: list[Path]) -> None:
        payload = {
            "fingerprint": file_fingerprint(files),
            "chunks": [chunk.__dict__ for chunk in self.chunks],
        }
        INDEX_PATH.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


knowledge_base = KnowledgeBase()


def find_knowledge_files() -> Iterable[Path]:
    if not KNOWLEDGE_DIR.exists():
        return []
    extensions = {".txt", ".md", ".pdf", ".docx", ".xlsx"}
    ignored_names = {"README.md", "urls.txt", "urls.example.txt"}
    return sorted(
        path
        for path in KNOWLEDGE_DIR.rglob("*")
        if path.suffix.lower() in extensions and path.name not in ignored_names
    )


def file_fingerprint(files: list[Path]) -> list[dict[str, int | str]]:
    fingerprint = [
        {
            "path": str(path.relative_to(ROOT)),
            "size": path.stat().st_size,
            "mtime": int(path.stat().st_mtime),
        }
        for path in files
    ]
    fingerprint.append({"urls": "\n".join(read_urls())})
    return fingerprint


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".pdf":
        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix == ".docx":
        doc = Document(str(path))
        return "\n".join(paragraph.text for paragraph in doc.paragraphs)
    if suffix == ".xlsx":
        return extract_xlsx_text(path)
    return ""


def extract_xlsx_text(path: Path) -> str:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    parts = []
    for sheet in workbook.worksheets:
        parts.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                parts.append(" | ".join(values))
    workbook.close()
    return "\n".join(parts)


def read_urls() -> list[str]:
    if not URLS_PATH.exists():
        return []
    urls = []
    for line in URLS_PATH.read_text(encoding="utf-8", errors="ignore").splitlines():
        clean = line.strip()
        if clean and not clean.startswith("#"):
            urls.append(clean)
    return urls


async def fetch_url_text(url: str) -> str:
    async with httpx.AsyncClient(follow_redirects=True, timeout=20) as http:
        response = await http.get(url, headers={"User-Agent": "AquaFarmEU knowledge bot/1.0"})
        response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    title = soup.title.get_text(" ", strip=True) if soup.title else url
    body = soup.get_text("\n", strip=True)
    return f"Page: {title}\nURL: {url}\n\n{body}"


def split_text(text: str, max_chars: int = 1600, overlap: int = 250) -> list[str]:
    clean = "\n".join(line.strip() for line in text.splitlines() if line.strip())
    if not clean:
        return []

    chunks = []
    start = 0
    while start < len(clean):
        end = min(start + max_chars, len(clean))
        chunks.append(clean[start:end])
        if end == len(clean):
            break
        start = end - overlap
        if start < 0:
            break
    return chunks


async def embed_texts(texts: list[str]) -> list[list[float]]:
    embeddings: list[list[float]] = []
    batch_size = 64
    for start in range(0, len(texts), batch_size):
        batch = texts[start : start + batch_size]
        response = await client.embeddings.create(model=OPENAI_EMBEDDING_MODEL, input=batch)
        embeddings.extend(item.embedding for item in response.data)
    return embeddings


def is_allowed(update: Update) -> bool:
    if not ALLOWED_USER_IDS:
        return True
    user = update.effective_user
    return bool(user and user.id in ALLOWED_USER_IDS)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id if update.effective_user else "unknown"
    await update.message.reply_text(
        "Hi. I am the AquaFarmEU project assistant.\n\n"
        f"Your Telegram user ID is: {user_id}\n\n"
        "Ask me a question about the project documents."
    )


async def reload_knowledge(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        await update.message.reply_text("Sorry, you do not have access to this bot.")
        return

    await update.message.reply_text("Reloading project knowledge...")
    await knowledge_base.load()
    await update.message.reply_text(f"Done. I loaded {len(knowledge_base.chunks)} knowledge chunks.")


async def answer_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        await update.message.reply_text("Sorry, you do not have access to this bot.")
        return

    question = update.message.text.strip()
    if not question:
        return

    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")

    if not knowledge_base.loaded:
        await update.message.reply_text("I am loading the project documents. This can take a minute on the first run.")
        try:
            await knowledge_base.load()
        except Exception:
            logger.exception("Knowledge index failed to build")
            await update.message.reply_text(
                "I could not load the project documents. Please check the Render logs."
            )
            return

    if not knowledge_base.chunks:
        await update.message.reply_text(
            "I do not have project documents loaded yet. Add files to the knowledge folder and restart me."
        )
        return

    query_embedding = (await embed_texts([question]))[0]
    relevant_chunks = knowledge_base.search(query_embedding)

    if not relevant_chunks:
        await update.message.reply_text(
            "I could not find this in the project documents. Try asking differently or add more source material."
        )
        return

    context_text = "\n\n".join(
        f"Source: {chunk.source}\n{chunk.text}" for chunk in relevant_chunks
    )

    response = await client.chat.completions.create(
        model=OPENAI_CHAT_MODEL,
        temperature=0.2,
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a helpful project assistant for AquaFarmEU. "
                    "Answer using only the supplied project context. "
                    "If the context is not enough, say that the documents do not contain the answer. "
                    "Keep answers clear, practical, and concise. "
                    "At the end, list source filenames you used."
                ),
            },
            {
                "role": "user",
                "content": f"Project context:\n{context_text}\n\nQuestion:\n{question}",
            },
        ],
    )

    answer = response.choices[0].message.content or "I could not prepare an answer."
    await update.message.reply_text(answer[:4000])


async def post_init(application: Application) -> None:
    logger.info("Bot started. Knowledge will be loaded on first question or /reload.")


def build_app() -> Application:
    app = Application.builder().token(TELEGRAM_BOT_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("reload", reload_knowledge))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, answer_question))
    return app


def main() -> None:
    app = build_app()

    if WEBHOOK_URL:
        logger.info("Starting webhook on port %s", PORT)
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            url_path="telegram-webhook",
            webhook_url=f"{WEBHOOK_URL}/telegram-webhook",
        )
    else:
        logger.info("Starting local polling")
        app.run_polling()


if __name__ == "__main__":
    main()
